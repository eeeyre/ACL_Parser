from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def generate_headers(table_type):
    if table_type == 'errors':
        header = ['Parsing Errors']
    if table_type == 'rules':
        header = ['ACL Name', 'Entry Number', 'Entry Type',
                  'Remark Text', 'Permission', 'Protocol',
                  'Source', 'Source Mask', 'Src Port Operator',
                  'Source Port', 'Source Port Range End',
                  'Destination IP', 'Destination Mask', 'Destination Port Operator',
                  'Destination Port', 'Destination Port Range End',
                  'Hit Count', 'Check Sum']
    if table_type == 'audit':
        header = ['ACL Name', 'Entry Number', 'Violation Type', 'Violation Description']
    return header


def get_sample_data(data_type):
    if data_type == 'errors':
        header = ['ERROR']
    if data_type == 'rules':
        header = ['AWS-FH-OUT', '1',
                  'extended', '', 'permit', 'icmp',
                  '0.0.0.0', '0.0.0.0', '', '', '',
                  '0.0.0.0', '0.0.0.0', '', '', '',
                  '4202115', '0xc880d763']
    if data_type == 'audit':
        header = ['AUDITLINE']
    return header


def output_xlsx_file(parsed_rules, errors, audit_type, audit, output_file):
    audit_bool = audit_type[0] or audit_type[1] or audit_type[2]
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Parsed'
    ws2 = wb.create_sheet('Errors')
    ws1.append(generate_headers('rules'))
    ws2.append(generate_headers('errors'))
    if audit_bool:
        ws3 = wb.create_sheet('Audit')
        ws3.append(generate_headers('audit'))
    parsed_count = 0
    err_count = 0
    audit_count = 0
    for acl in parsed_rules:
        for entry in acl:
            ws1.append(entry)
            parsed_count += 1
    for acl_err in errors:
        for entry_err in acl_err:
            ws2.append([entry_err])
            err_count += 1
    if audit_bool:
        for entry_audit in audit:
            ws3.append(entry_audit)
            audit_count += 1

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    parsed_range = "A1:R"+str(parsed_count+1)
    errors_range = "A1:A"+str(err_count+1)
    if audit_bool:
        audit_range = "A1:D"+str(audit_count+1)
    parsed_tab = Table(displayName='Parsed_Rules', ref=parsed_range, tableStyleInfo=style)
    errors_tab = Table(displayName='Parsing_Errors', ref=errors_range, tableStyleInfo=style)
    if audit_bool:
        audit_tab = Table(displayName='Auditing_Results', ref=audit_range, tableStyleInfo=style)

    ws1.add_table(parsed_tab)
    ws2.add_table(errors_tab)
    if audit_bool:
        ws3.add_table(audit_tab)

    wb.save(output_file)


def add_desc(desc, addition):
    output = ''
    if desc != '':
        output = desc + ", " + addition
    else:
        output = addition
    return output
